from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side
import os
import re
from openpyxl.styles import Alignment
import logging

logging.getLogger().handlers.clear()
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("search_log.log", mode='a', encoding='utf-8'),  # Nastavení UTF-8
    ]
)


class ExcelWriter:
    def __init__(self, output_folder="OUTPUT"):
        self.output_folder = output_folder
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder)

    def create_output_file_with_search_query(self, search_query, filename_search_query):
        """
        Creates an Excel file containing the search query in a formatted cell.

        :param search_query: The search query string to include in the file.
        :param filename_search_query: The name of the file (without extension) to be created.
        :return: The file path of the created Excel file, or `None` if an error occurs.
        """
        try:
            file_name = f"{filename_search_query}.xlsx"
            file_path = os.path.join(self.output_folder, file_name)

            workbook = Workbook()
            sheet = workbook.active

            cell = sheet["B2"]
            cell.value = f"Search query: {search_query}"
            
            cell.font = Font(bold=True, size=12)

            sheet.column_dimensions["A"].width = 4
            sheet.column_dimensions["B"].width = 120

            border = Border(
                left=Side(style="thick"),
                right=Side(style="thick"),
                top=Side(style="thick"),
                bottom=Side(style="thick")
            )
            sheet["B2"].border = border

            sheet["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            workbook.save(file_path)
            logging.info(f"File created successfully: {file_path}")
            return file_path

        except Exception as e:
            logging.error(f"An error occurred while creating the file: {str(e)}")
            print(f"An error occurred while creating the file: {str(e)}")
            return None
        
    def add_urls_to_output_file(self, file_path, url):
        """
        Adds a URL to an existing Excel file in a new row with a clickable hyperlink.

        :param file_path: The path to the Excel file where the URL will be added.
        :param url: The URL to insert into the file.
        """
        try:
            if not os.path.exists(file_path):
                logging.error(f"The file {file_path} does not exist. Please create the file first.")
                print(f"The file {file_path} does not exist. Please create the file first.")
                return

            workbook = load_workbook(file_path)
            sheet = workbook.active

            last_row = sheet.max_row + 1

            sheet[f"B{last_row}"] = url
            sheet[f"B{last_row}"].hyperlink = url  
            sheet[f"B{last_row}"].style = "Hyperlink"  

            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            sheet[f"B{last_row}"].border = border

            workbook.save(file_path)
            logging.info(f"URL '{url}' successfully added to file: {file_path}")

        except Exception as e:
            logging.error(f"An error occurred while writing URL to the file: {str(e)}")
            print(f"An error occurred while writing URL to the file: {str(e)}")

    def modify_serach_query_for_filename(self, query):
        """
        Sanitizes a search query string to make it suitable for use as a filename.

        :param query: The original search query string.
        :return: A sanitized version of the query with:
            - Non-alphanumeric characters removed (except spaces and hyphens).
            - Leading and trailing whitespace trimmed.
        """
        modified_query = re.sub(r'[^\w\s-]', '', query)  # Removes illegal characters
        modified_query = modified_query.strip()  # Remove leading and trailing spaces
        return modified_query

